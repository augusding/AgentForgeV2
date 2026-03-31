"""AgentForge V2 — 文件上传路由"""

import json
import logging
import os
from uuid import uuid4

from aiohttp import web

logger = logging.getLogger(__name__)

_MAX_FILE_SIZE = int(os.environ.get("MAX_UPLOAD_SIZE_MB", "20")) * 1024 * 1024

_ALLOWED_EXTENSIONS = {
    ".pdf", ".docx", ".doc", ".txt", ".md", ".csv", ".json",
    ".xlsx", ".xls", ".pptx", ".ppt",
    ".png", ".jpg", ".jpeg", ".gif", ".webp", ".zip",
}
_MAX_FILENAME_LENGTH = 255
_MAGIC_SIGNATURES: dict[str, list[bytes]] = {
    ".pdf": [b"%PDF"],
    ".docx": [b"PK\x03\x04"],
    ".xlsx": [b"PK\x03\x04"],
    ".pptx": [b"PK\x03\x04"],
    ".zip": [b"PK\x03\x04"],
    ".png": [b"\x89PNG"],
    ".jpg": [b"\xff\xd8\xff"],
    ".jpeg": [b"\xff\xd8\xff"],
    ".gif": [b"GIF87a", b"GIF89a"],
}


def _check_file_type(filename: str, file_path) -> str | None:
    """检查文件类型安全性。返回错误信息或 None。"""
    from pathlib import Path
    ext = Path(filename).suffix.lower()
    if ext not in _ALLOWED_EXTENSIONS:
        return f"不支持的文件类型: {ext}"
    if len(filename) > _MAX_FILENAME_LENGTH:
        return f"文件名过长（最大 {_MAX_FILENAME_LENGTH} 字符）"
    if ext in _MAGIC_SIGNATURES and Path(file_path).exists():
        with open(file_path, "rb") as f:
            header = f.read(8)
        if not any(header.startswith(sig) for sig in _MAGIC_SIGNATURES[ext]):
            return f"文件内容与后缀 {ext} 不匹配（可能被篡改）"
    return None


def _json(data, status=200):
    return web.Response(
        text=json.dumps(data, ensure_ascii=False, default=str),
        content_type="application/json", status=status,
    )


async def handle_upload(request: web.Request) -> web.Response:
    """
    POST /api/v1/files/upload (multipart/form-data)
    字段: file (文件), target ("knowledge" | "chat"), doc_id (可选)
    """
    engine = request.app["engine"]
    user = request.get("user") or {}
    u_id = user.get("sub", "anonymous") if isinstance(user, dict) else "anonymous"
    o_id = user.get("org_id", "_default") if isinstance(user, dict) else "_default"
    upload_dir = engine.root_dir / "data" / "uploads" / (o_id or "_default") / u_id
    upload_dir.mkdir(parents=True, exist_ok=True)

    reader = await request.multipart()
    file_field = None
    file_path = None
    target = "chat"
    doc_id = ""

    async for field in reader:
        if field.name == "file":
            file_field = field
            filename = field.filename or f"upload_{uuid4().hex[:8]}"
            filename = filename.replace("/", "_").replace("\\", "_").replace("..", "_")
            file_path = upload_dir / f"{uuid4().hex[:8]}_{filename}"
            size = 0
            with open(file_path, "wb") as f:
                while True:
                    chunk = await field.read_chunk()
                    if not chunk:
                        break
                    size += len(chunk)
                    if size > _MAX_FILE_SIZE:
                        os.unlink(file_path)
                        return _json({"error": f"文件超过大小限制 ({_MAX_FILE_SIZE // 1024 // 1024}MB)"}, 413)
                    f.write(chunk)
        elif field.name == "target":
            target = (await field.read()).decode("utf-8", errors="replace")
        elif field.name == "doc_id":
            doc_id = (await field.read()).decode("utf-8", errors="replace")

    if not file_field or not file_path or not file_path.exists():
        return _json({"error": "未收到文件"}, status=400)

    # 文件类型安全检查
    type_error = _check_file_type(file_field.filename or "", file_path)
    if type_error:
        try:
            os.unlink(file_path)
        except Exception:
            pass
        return _json({"error": type_error}, status=400)

    # 单用户存储限额检查
    max_storage = int(os.environ.get("MAX_USER_STORAGE_MB", "500")) * 1024 * 1024
    if upload_dir.exists():
        current_size = sum(f.stat().st_size for f in upload_dir.rglob("*") if f.is_file())
        if current_size > max_storage:
            try:
                os.unlink(file_path)
            except Exception:
                pass
            return _json({"error": f"存储空间已满（已用 {current_size // 1024 // 1024}MB / {max_storage // 1024 // 1024}MB）"}, status=413)

    from core.file_parser import extract_text
    extracted = await extract_text(str(file_path))

    try:
        rel_path = str(file_path.relative_to(engine.root_dir)).replace("\\", "/")
    except ValueError:
        rel_path = str(file_path).replace("\\", "/")
    result = {
        "file_id": file_path.name,
        "filename": file_field.filename,
        "size": file_path.stat().st_size,
        "path": rel_path,
        "extracted_length": len(extracted),
    }

    if target == "knowledge" and extracted and engine.knowledge_base:
        # 将文件移到知识库专用目录
        kb_dir = engine.root_dir / "data" / "knowledge" / (o_id or "_default") / u_id
        kb_dir.mkdir(parents=True, exist_ok=True)
        kb_path = kb_dir / file_path.name
        import shutil
        shutil.move(str(file_path), str(kb_path))
        file_path = kb_path
        result["path"] = str(kb_path)

        final_doc_id = doc_id or file_path.stem
        chunks = engine.knowledge_base.add_document(
            doc_id=final_doc_id,
            content=extracted,
            metadata={"source": "upload", "source_type": "upload", "filename": file_field.filename},
            is_markdown=file_path.suffix.lower() == ".md",
            org_id=o_id,
            user_id=u_id,
        )
        result["knowledge"] = {"doc_id": final_doc_id, "chunks": chunks}

    if target == "chat":
        from core.media_processor import is_audio, transcribe_audio
        from pathlib import Path as _P
        if is_audio(_P(str(file_path))):
            stt_text = await transcribe_audio(_P(str(file_path)))
            result["extracted_text"] = stt_text
            result["media_type"] = "audio"
        else:
            result["extracted_text"] = extracted[:3000]

    return _json(result)


async def handle_list_files(request: web.Request) -> web.Response:
    """GET /api/v1/files — 列出已上传的文件"""
    engine = request.app["engine"]
    user = request.get("user") or {}
    u_id = user.get("sub", "anonymous") if isinstance(user, dict) else "anonymous"
    o_id = user.get("org_id", "_default") if isinstance(user, dict) else "_default"
    upload_dir = engine.root_dir / "data" / "uploads" / (o_id or "_default") / u_id
    if not upload_dir.exists():
        return _json({"files": []})

    files = []
    for f in sorted(upload_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if f.is_file():
            files.append({
                "file_id": f.stem,
                "filename": f.name,
                "size": f.stat().st_size,
                "modified": f.stat().st_mtime,
            })
    return _json({"files": files[:100]})


async def handle_download(request: web.Request) -> web.Response:
    """GET /api/v1/files/download/{path:.*} — 下载文件"""
    raw_path = request.match_info.get("path", "")
    engine = request.app["engine"]
    file_path = (engine.root_dir / raw_path).resolve()
    if not str(file_path).startswith(str(engine.root_dir)):
        return web.Response(status=403, text="路径越界")
    if not file_path.is_file():
        return web.Response(status=404, text="文件不存在")

    _MIME = {
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".doc": "application/msword",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        ".ppt": "application/vnd.ms-powerpoint",
        ".pdf": "application/pdf",
        ".csv": "text/csv; charset=utf-8",
        ".txt": "text/plain; charset=utf-8",
        ".md": "text/markdown; charset=utf-8",
        ".json": "application/json; charset=utf-8",
        ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".gif": "image/gif", ".webp": "image/webp",
        ".zip": "application/zip",
    }
    from urllib.parse import quote
    ct = _MIME.get(file_path.suffix.lower(), "application/octet-stream")
    encoded = quote(file_path.name, safe='')
    return web.FileResponse(file_path, headers={
        "Content-Type": ct,
        "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}",
    })


async def handle_preview(request: web.Request) -> web.Response:
    """GET /api/v1/files/preview/{path:.*} — 按格式返回结构化预览"""
    raw = request.match_info.get("path", "")
    engine = request.app["engine"]
    fp = (engine.root_dir / raw).resolve()
    if not str(fp).startswith(str(engine.root_dir)):
        return _json({"error": "路径越界"}, 403)
    if not fp.is_file():
        return _json({"error": "文件不存在"}, 404)
    suffix = fp.suffix.lower()
    fmt = suffix.lstrip(".")
    base = {"filename": fp.name, "size": fp.stat().st_size, "format": fmt}

    if suffix in (".xlsx", ".xls"):
        return _json({**base, "type": "table", "data": _preview_xlsx(fp)})
    if suffix == ".csv":
        return _json({**base, "type": "table", "data": _preview_csv(fp)})
    if suffix == ".pptx":
        return _json({**base, "type": "slides", "data": _preview_pptx(fp)})
    if suffix in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
        return _json({**base, "type": "image", "data": {"url": f"/api/v1/files/download/{raw}"}})
    if suffix == ".md":
        return _json({**base, "type": "markdown", "content": fp.read_text(encoding="utf-8", errors="replace")[:10000]})

    from core.file_parser import extract_text
    text = await extract_text(str(fp))
    ptype = "richtext" if suffix in (".docx", ".pdf") else "text"
    return _json({**base, "type": ptype, "content": text[:10000] if text else ""})


def _preview_xlsx(fp):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames[:5]:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(max_row=200, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append(cells)
            if rows:
                sheets.append({"name": name, "rows": rows, "total_rows": ws.max_row or len(rows)})
        wb.close()
        return {"sheets": sheets}
    except Exception as e:
        return {"error": str(e)}


def _preview_csv(fp):
    import csv
    rows = []
    try:
        with open(fp, "r", encoding="utf-8-sig", errors="replace") as f:
            for i, row in enumerate(csv.reader(f)):
                if i >= 200: break
                rows.append(row)
        return {"sheets": [{"name": "数据", "rows": rows, "total_rows": len(rows)}]}
    except Exception as e:
        return {"error": str(e)}


def _preview_pptx(fp):
    """PPT → 幻灯片列表。直接读 ZIP 内 XML，绕过 python-pptx rId bug。"""
    import zipfile, re
    try:
        slides = []
        with zipfile.ZipFile(str(fp)) as z:
            slide_files = sorted(
                [n for n in z.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")],
                key=lambda n: int(m.group(1)) if (m := re.search(r'slide(\d+)', n)) else 0,
            )
            for idx, sf in enumerate(slide_files[:30], 1):
                xml = z.read(sf).decode("utf-8", errors="replace")
                texts = [t.strip() for t in re.findall(r'<a:t>([^<]+)</a:t>', xml) if t.strip()]
                title = texts[0] if texts else f"幻灯片 {idx}"
                content = texts[1:] if len(texts) > 1 else []
                slides.append({"number": idx, "title": title, "content": content})
        return {"slides": slides, "total": len(slides)}
    except Exception as e:
        return {"error": str(e)}


async def handle_serve(request: web.Request) -> web.Response:
    """GET /api/v1/files/serve/{path:.*} — 内联展示文件（不触发下载）"""
    engine = request.app["engine"]
    raw_path = request.match_info["path"]
    from pathlib import Path
    candidates = [engine.root_dir / raw_path,
                  engine.root_dir / "data" / "uploads" / Path(raw_path).name,
                  engine.root_dir / "data" / "outputs" / Path(raw_path).name]
    file_path = None
    for c in candidates:
        try:
            resolved = c.resolve()
            if resolved.is_file() and str(resolved).startswith(str(engine.root_dir.resolve())):
                file_path = resolved; break
        except Exception:
            continue
    if not file_path:
        return web.json_response({"error": f"文件不存在: {raw_path}"}, status=404)
    ct_map = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
              ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp",
              ".pdf": "application/pdf"}
    ct = ct_map.get(file_path.suffix.lower(), "application/octet-stream")
    return web.FileResponse(path=file_path, headers={
        "Content-Disposition": f'inline; filename="{file_path.name}"',
        "Content-Type": ct, "Cache-Control": "private, max-age=3600"})


def register(app: web.Application) -> None:
    app.router.add_post("/api/v1/files/upload", handle_upload)
    app.router.add_get("/api/v1/files", handle_list_files)
    app.router.add_get("/api/v1/files/preview/{path:.*}", handle_preview)
    app.router.add_get("/api/v1/files/download/{path:.*}", handle_download)
    app.router.add_get("/api/v1/files/serve/{path:.*}", handle_serve)
